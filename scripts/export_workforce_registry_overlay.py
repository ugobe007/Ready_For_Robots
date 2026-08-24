#!/usr/bin/env python3
"""Export the ChatGPT workforce-registry workbook to a researcher overlay JSON.

This is NOT a catalog compiler. Product names stay `researcher_claim`.
Never write these names into primary_robots, FIND indexes, or
robot_employment_universe_v1.json without an OEM-page check.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "calibration" / "readyforrobots_robot_workforce_registry_v1.xlsx"
OUT = ROOT / "docs" / "calibration" / "robot_workforce_registry_overlay_v1.json"

HEADER_MAP = {
    "ID": "id",
    "Company": "company",
    "Robot / Product": "claimed_product",
    "Primary Employment Category": "primary_category",
    "Additional Categories": "additional_categories",
    "Robot Class": "robot_class",
    "Work Families / Jobs": "work_families",
    "Core Capabilities": "core_capabilities",
    "Target Industries": "target_industries",
    "Commercialization Stage": "commercialization_stage",
    "Readiness Score (1-5)": "readiness_score",
    "R4R Placement Priority": "placement_priority",
    "Available for Work?": "available_for_work",
    "Primary Geography": "geography",
    "Source Type": "source_type",
    "Source URL": "source_url",
}


def export(xlsx: Path = XLSX, out: Path = OUT) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Robot Workforce Registry"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    unknown = [h for h in headers if h not in HEADER_MAP]
    if unknown:
        raise SystemExit(f"unexpected registry headers: {unknown}")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        rec: dict = {}
        for header, value in zip(headers, raw):
            key = HEADER_MAP[header]
            if isinstance(value, str):
                value = value.strip() or None
            rec[key] = value
        if not rec.get("company"):
            continue
        rec["epistemic"] = "researcher_claim"
        rec["do_not_treat_as_catalog"] = True
        rows.append(rec)

    payload = {
        "version": "1.0",
        "kind": "researcher_overlay",
        "source_file": "docs/calibration/readyforrobots_robot_workforce_registry_v1.xlsx",
        "warning": (
            "One row per company. Product names are researcher claims, not "
            "OEM-page catalog. Do not write into primary_robots or FIND "
            "indexes until verified on the source_url."
        ),
        "shape": {
            "rows": "one_company_one_product_cell",
            "product_cap": "not_enforced",
            "official_source_stamp": "rubber_stamped_on_all_rows",
        },
        "dashboard_claims": {
            "companies": 200,
            "high_priority": 172,
            "available_for_work_now": 172,
            "rfr_verdict": "too_optimistic_not_placement_ready",
        },
        "n": len(rows),
        "rows": rows,
    }
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return payload


def main() -> int:
    payload = export()
    print(f"wrote {OUT.relative_to(ROOT)} n={payload['n']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
